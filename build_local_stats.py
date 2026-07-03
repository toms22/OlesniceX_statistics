from __future__ import annotations

import csv
import json
import math
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_SOURCE = ROOT / "Zdroj_dat" / "Výsledky Olešnice X final.xlsx"
OUT_DIR = ROOT / "vystupy"


def find_source_workbook() -> Path:
    if DEFAULT_SOURCE.exists():
        return DEFAULT_SOURCE
    candidates = sorted((ROOT / "Zdroj_dat").glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"Ve složce {ROOT / 'Zdroj_dat'} není žádný .xlsx soubor.")
    return candidates[0]


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    if text.upper() in {"", "NP", "DNS", "DSQ", "MS", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(value) if value.is_integer() else None
    text = str(value).strip()
    if not text:
        return None
    number = parse_float(text)
    if number is not None and number.is_integer():
        return int(number)
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def split_leagues(value: Any) -> list[str]:
    if value is None or str(value).strip() == "":
        return ["Bez ligy"]
    parts = []
    for part in str(value).replace(";", ",").split(","):
        clean = part.strip()
        if not clean:
            continue
        normalized = clean.upper()
        if normalized == "FSEU":
            continue
        if normalized in {"BIOMAC HE", "HE"}:
            clean = "HE"
        parts.append(clean)
    return parts or ["Bez ligy"]


def fmt_num(value: float | int | None, decimals: int = 2) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}".replace(".", ",")


def fmt_date(value: date | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


@dataclass
class Run:
    season: int
    date: date | None
    place: str
    leagues: list[str]
    league_text: str
    l: float | None
    p: float | None
    result: float | None
    result_raw: Any
    hose: str
    placement: int | None


def read_runs(source: Path | None = None) -> list[Run]:
    wb = load_workbook(source or find_source_workbook(), data_only=True, read_only=True)
    runs: list[Run] = []
    for sheet in wb.sheetnames:
        if not sheet.isdigit():
            continue
        season = int(sheet)
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 9:
                continue
            dt = parse_date(row[1])
            place = str(row[2]).strip() if row[2] is not None else ""
            if not dt and not place:
                continue
            hose = str(row[7]).strip().upper() if row[7] is not None else ""
            if hose not in {"2B", "3B"}:
                hose = hose or "Neznámé"
            runs.append(
                Run(
                    season=season,
                    date=dt,
                    place=place,
                    leagues=split_leagues(row[3]),
                    league_text=", ".join(split_leagues(row[3])),
                    l=parse_float(row[4]),
                    p=parse_float(row[5]),
                    result=parse_float(row[6]),
                    result_raw=row[6],
                    hose=hose,
                    placement=parse_int(row[8]),
                )
            )
    wb.close()
    return runs


def mean(values: Iterable[float]) -> float | None:
    xs = list(values)
    return statistics.fmean(xs) if xs else None


def median(values: Iterable[float]) -> float | None:
    xs = list(values)
    return statistics.median(xs) if xs else None


def best_run(rows: list[Run], attr: str) -> Run | None:
    valid = [r for r in rows if getattr(r, attr) is not None]
    return min(valid, key=lambda r: getattr(r, attr)) if valid else None


def best_target(rows: list[Run]) -> tuple[float | None, Run | None, str]:
    best: tuple[float, Run, str] | None = None
    for run in rows:
        for side, value in (("L", run.l), ("P", run.p)):
            if value is None:
                continue
            if best is None or value < best[0]:
                best = (value, run, side)
    if not best:
        return None, None, ""
    return best


def stats_for(rows: list[Run]) -> dict[str, Any]:
    result_values = [r.result for r in rows if r.result is not None]
    l_values = [r.l for r in rows if r.l is not None]
    p_values = [r.p for r in rows if r.p is not None]
    all_targets = l_values + p_values
    best_result = best_run(rows, "result")
    best_l = best_run(rows, "l")
    best_p = best_run(rows, "p")
    best_t, best_t_run, best_t_side = best_target(rows)
    placements = Counter(r.placement if r.placement is not None else "Bez pořadí" for r in rows)
    return {
        "pocet_zavodu": len(rows),
        "pocet_vysledku_do_statistik": len(result_values),
        "pocet_np_bez_vysledku": len(rows) - len(result_values),
        "nejlepsi_cas": min(result_values) if result_values else None,
        "nejlepsi_cas_detail": best_result,
        "prumer_vysledku": mean(result_values),
        "median_vysledku": median(result_values),
        "nejlepsi_l": min(l_values) if l_values else None,
        "nejlepsi_l_detail": best_l,
        "prumer_l": mean(l_values),
        "median_l": median(l_values),
        "nejlepsi_p": min(p_values) if p_values else None,
        "nejlepsi_p_detail": best_p,
        "prumer_p": mean(p_values),
        "median_p": median(p_values),
        "nejlepsi_sestrik": best_t,
        "nejlepsi_sestrik_detail": best_t_run,
        "nejlepsi_sestrik_strana": best_t_side,
        "prumer_sestriku": mean(all_targets),
        "median_sestriku": median(all_targets),
        "umisteni": placements,
    }


def detail(run: Run | None, attr: str | None = None) -> str:
    if not run:
        return ""
    value = getattr(run, attr) if attr else run.result
    return f"{fmt_num(value)} ({fmt_date(run.date)}, {run.place}, {run.hose})"


def placement_text(counter: Counter) -> str:
    def key(item: tuple[Any, int]) -> tuple[int, Any]:
        pos, _ = item
        return (999999, pos) if isinstance(pos, str) else (pos, "")

    return ", ".join(f"{pos}: {count}x" for pos, count in sorted(counter.items(), key=key))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def season_sort_value(season: int | str) -> int | None:
    return season if isinstance(season, int) else None


def metric_row(label: str, rows: list[Run], season: int | str, hose: str | None = None) -> dict[str, Any]:
    s = stats_for(rows)
    return {
        "sezona": season,
        "sezona_sort": season_sort_value(season),
        "hadice": hose or "Celkem",
        "skupina": label,
        "pocet_zavodu": s["pocet_zavodu"],
        "pocet_vysledku_do_statistik": s["pocet_vysledku_do_statistik"],
        "pocet_np_bez_vysledku": s["pocet_np_bez_vysledku"],
        "nejlepsi_cas": s["nejlepsi_cas"],
        "nejlepsi_cas_detail": detail(s["nejlepsi_cas_detail"], "result"),
        "prumer_vysledku": s["prumer_vysledku"],
        "median_vysledku": s["median_vysledku"],
        "nejlepsi_l": s["nejlepsi_l"],
        "nejlepsi_l_detail": detail(s["nejlepsi_l_detail"], "l"),
        "prumer_l": s["prumer_l"],
        "median_l": s["median_l"],
        "nejlepsi_p": s["nejlepsi_p"],
        "nejlepsi_p_detail": detail(s["nejlepsi_p_detail"], "p"),
        "prumer_p": s["prumer_p"],
        "median_p": s["median_p"],
        "nejlepsi_sestrik": s["nejlepsi_sestrik"],
        "nejlepsi_sestrik_detail": (
            f"{fmt_num(s['nejlepsi_sestrik'])} {s['nejlepsi_sestrik_strana']} "
            f"({fmt_date(s['nejlepsi_sestrik_detail'].date)}, {s['nejlepsi_sestrik_detail'].place}, {s['nejlepsi_sestrik_detail'].hose})"
            if s["nejlepsi_sestrik_detail"]
            else ""
        ),
        "prumer_sestriku": s["prumer_sestriku"],
        "median_sestriku": s["median_sestriku"],
        "umisteni": placement_text(s["umisteni"]),
    }


def blank_time_metrics(row: dict[str, Any]) -> None:
    for key in (
        "nejlepsi_cas",
        "nejlepsi_cas_detail",
        "prumer_vysledku",
        "median_vysledku",
        "nejlepsi_l",
        "nejlepsi_l_detail",
        "prumer_l",
        "median_l",
        "nejlepsi_p",
        "nejlepsi_p_detail",
        "prumer_p",
        "median_p",
        "nejlepsi_sestrik",
        "nejlepsi_sestrik_detail",
        "prumer_sestriku",
        "median_sestriku",
    ):
        row[key] = None if not key.endswith("_detail") else ""


def main() -> int:
    OUT_DIR.mkdir(exist_ok=True)
    source = find_source_workbook()
    runs = read_runs(source)
    seasons = sorted({r.season for r in runs})

    raw_rows = []
    for r in runs:
        raw_rows.append(
            {
                "sezona": r.season,
                "sezona_sort": r.season,
                "datum": fmt_date(r.date),
                "misto": r.place,
                "liga": r.league_text,
                "L": r.l,
                "P": r.p,
                "vysledny_cas": r.result,
                "vysledny_cas_raw": r.result_raw,
                "hadice": r.hose,
                "poradi": r.placement,
            }
        )
    write_csv(
        OUT_DIR / "data_export.csv",
        raw_rows,
        ["sezona", "sezona_sort", "datum", "misto", "liga", "L", "P", "vysledny_cas", "vysledny_cas_raw", "hadice", "poradi"],
    )

    season_rows: list[dict[str, Any]] = []
    for season in seasons:
        season_runs = [r for r in runs if r.season == season]
        season_rows.append(metric_row("Sezona celkem", season_runs, season))
        for hose in ("2B", "3B"):
            subset = [r for r in season_runs if r.hose == hose]
            if subset:
                season_rows.append(metric_row(f"Sezona {hose}", subset, season, hose))
    metric_fields = list(season_rows[0].keys())
    write_csv(OUT_DIR / "statistiky_sezony.csv", season_rows, metric_fields)

    overall_rows = [metric_row("Všechny sezony", runs, "2009-2026")]
    for hose in ("2B", "3B"):
        overall_rows.append(metric_row(f"Všechny sezony {hose}", [r for r in runs if r.hose == hose], "2009-2026", hose))
    write_csv(OUT_DIR / "statistiky_celkem.csv", overall_rows, metric_fields)

    league_counter: dict[tuple[str, str], list[Run]] = defaultdict(list)
    for r in runs:
        for league in r.leagues:
            league_counter[(league, r.hose)].append(r)
            league_counter[(league, "Celkem")].append(r)
    league_rows = []
    for (league, hose), subset in sorted(league_counter.items(), key=lambda x: (x[0][0], x[0][1])):
        row = metric_row(league, subset, "2009-2026", hose)
        row["liga"] = league
        league_rows.append(row)
    write_csv(OUT_DIR / "statistiky_ligy.csv", league_rows, ["liga"] + metric_fields)

    event_counter: dict[tuple[str, str], list[Run]] = defaultdict(list)
    event_totals: dict[str, list[Run]] = defaultdict(list)
    for r in runs:
        event_totals[r.place].append(r)
        event_counter[(r.place, r.hose)].append(r)
        event_counter[(r.place, "Celkem")].append(r)
    event_rows = []
    for place, total_subset in sorted(event_totals.items(), key=lambda item: (-len(item[1]), item[0])):
        grouped_hoses = sorted({r.hose for r in total_subset if r.hose in {"2B", "3B"}})
        row = metric_row(place, event_counter[(place, "Celkem")], "2009-2026", "Celkem")
        row["soutez"] = place
        row["ma_2b_i_3b"] = "ano" if {"2B", "3B"}.issubset(set(grouped_hoses)) else "ne"
        row["uroven"] = "Celkem"
        if row["ma_2b_i_3b"] == "ano":
            blank_time_metrics(row)
        event_rows.append(row)
        for hose in ("2B", "3B"):
            subset = event_counter.get((place, hose), [])
            if not subset:
                continue
            row = metric_row(place, subset, "2009-2026", hose)
            row["soutez"] = place
            row["ma_2b_i_3b"] = "ano" if {"2B", "3B"}.issubset(set(grouped_hoses)) else "ne"
            row["uroven"] = hose
            event_rows.append(row)
    write_csv(OUT_DIR / "statistiky_souteze.csv", event_rows, ["soutez", "uroven", "ma_2b_i_3b"] + metric_fields)

    placement_rows = []
    groups: dict[tuple[str, str, str], list[Run]] = defaultdict(list)
    for r in runs:
        pos = str(r.placement) if r.placement is not None else "Bez pořadí"
        groups[("Celkem", r.hose, pos)].append(r)
        groups[("Celkem", "Celkem", pos)].append(r)
        for league in r.leagues:
            groups[(league, r.hose, pos)].append(r)
            groups[(league, "Celkem", pos)].append(r)
    for (group, hose, pos), subset in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1], int(x[0][2]) if x[0][2].isdigit() else 999999)):
        placement_rows.append({"skupina": group, "hadice": hose, "poradi": pos, "pocet": len(subset)})
    write_csv(OUT_DIR / "pocty_umisteni.csv", placement_rows, ["skupina", "hadice", "poradi", "pocet"])

    top_results_rows = []
    result_runs = sorted(
        runs,
        key=lambda r: (
            r.result is None,
            r.result if r.result is not None else 999999,
            r.date,
            r.place,
            r.hose,
        ),
    )
    for idx, run in enumerate(result_runs, start=1):
        top_results_rows.append(
            {
                "rank": idx,
                "hadice": run.hose,
                "cas": run.result,
                "cas_raw": run.result_raw,
                "datum": fmt_date(run.date),
                "sezona": run.season,
                "sezona_sort": run.season,
                "soutez": run.place,
                "liga": run.league_text,
                "L": run.l,
                "P": run.p,
                "poradi": run.placement,
            }
        )
    write_csv(OUT_DIR / "top_vysledky.csv", top_results_rows, ["rank", "hadice", "cas", "cas_raw", "datum", "sezona", "sezona_sort", "soutez", "liga", "L", "P", "poradi"])

    top_target_rows = []
    target_items = []
    for run in runs:
        if run.l is not None:
            target_items.append((run.l, "L", run))
        if run.p is not None:
            target_items.append((run.p, "P", run))
    for idx, (value, side, run) in enumerate(sorted(target_items, key=lambda item: item[0]), start=1):
        top_target_rows.append(
            {
                "rank": idx,
                "hadice": run.hose,
                "strana": side,
                "sestrik": value,
                "datum": fmt_date(run.date),
                "sezona": run.season,
                "sezona_sort": run.season,
                "soutez": run.place,
                "liga": run.league_text,
            }
        )
    write_csv(OUT_DIR / "top_sestriky.csv", top_target_rows, ["rank", "hadice", "strana", "sestrik", "datum", "sezona", "sezona_sort", "soutez", "liga"])

    report_lines = [
        "# Statistiky Olešnice X",
        "",
        f"Zdroj: `{source.name}`",
        f"Počet záznamů: {len(runs)}",
        "",
        "Poznámka: `NP`, `MS`, `DNS`, `DSQ` a prázdné výsledné časy nejsou započtené do statistik výsledného času. Pokud má řádek platný sestřik `L` nebo `P`, počítá se do statistik sestřiků i při `NP` výsledku.",
        "",
        "## Celkový souhrn",
        "",
        "| Skupina | Závody | Výsledky v metrikách | NP/bez výsledku | Nejlepší čas | Průměr výsledku | Medián výsledku | Nejlepší L | Nejlepší P | Nejlepší sestřik | Průměr sestřiků | Medián sestřiků |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in overall_rows:
        report_lines.append(
            f"| {row['hadice']} | {row['pocet_zavodu']} | {row['pocet_vysledku_do_statistik']} | {row['pocet_np_bez_vysledku']} | "
            f"{fmt_num(row['nejlepsi_cas'])} | {fmt_num(row['prumer_vysledku'])} | {fmt_num(row['median_vysledku'])} | "
            f"{fmt_num(row['nejlepsi_l'])} | {fmt_num(row['nejlepsi_p'])} | {fmt_num(row['nejlepsi_sestrik'])} | "
            f"{fmt_num(row['prumer_sestriku'])} | {fmt_num(row['median_sestriku'])} |"
        )

    report_lines += ["", "## Po sezonách", ""]
    for season in seasons:
        report_lines += [f"### {season}", ""]
        rows = [row for row in season_rows if row["sezona"] == season]
        report_lines.append("| Hadice | Závody | Nejlepší čas | Průměr výsledku | Medián výsledku | Nejlepší L | Průměr L | Medián L | Nejlepší P | Průměr P | Medián P | Umístění |")
        report_lines.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---|")
        for row in rows:
            report_lines.append(
                f"| {row['hadice']} | {row['pocet_zavodu']} | {fmt_num(row['nejlepsi_cas'])} | "
                f"{fmt_num(row['prumer_vysledku'])} | {fmt_num(row['median_vysledku'])} | "
                f"{fmt_num(row['nejlepsi_l'])} | {fmt_num(row['prumer_l'])} | {fmt_num(row['median_l'])} | "
                f"{fmt_num(row['nejlepsi_p'])} | {fmt_num(row['prumer_p'])} | {fmt_num(row['median_p'])} | "
                f"{row['umisteni']} |"
            )
        report_lines.append("")

    report_lines += [
        "## Nejčastější soutěže",
        "",
        "| Soutěž | Rozdělení | Počet | Nejlepší čas | Průměr výsledku | Medián výsledku | Nejlepší sestřik |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    top_places = [row["soutez"] for row in event_rows if row["uroven"] == "Celkem"][:30]
    for place in top_places:
        place_rows = [row for row in event_rows if row["soutez"] == place]
        total_row = next(row for row in place_rows if row["uroven"] == "Celkem")
        rows_to_show = [total_row]
        if total_row["ma_2b_i_3b"] == "ano":
            rows_to_show.extend(row for row in place_rows if row["uroven"] in {"2B", "3B"})
        for row in rows_to_show:
            display_place = row["soutez"] if row["uroven"] == "Celkem" else ""
            split_label = row["uroven"]
            if row["uroven"] == "Celkem" and row["ma_2b_i_3b"] == "ano":
                split_label = "Celkem, časy níže 2B/3B"
            report_lines.append(
                f"| {display_place} | {split_label} | {row['pocet_zavodu']} | {fmt_num(row['nejlepsi_cas'])} | "
                f"{fmt_num(row['prumer_vysledku'])} | {fmt_num(row['median_vysledku'])} | {fmt_num(row['nejlepsi_sestrik'])} |"
            )

    report_lines += [
        "",
        "## Top výsledné časy",
        "",
        "| Hadice | Čas | Datum | Soutěž | Liga | L | P | Pořadí |",
        "|---|---:|---|---|---|---:|---:|---:|",
    ]
    for row in top_results_rows[:10] + top_results_rows[25:35]:
        report_lines.append(
            f"| {row['hadice']} | {fmt_num(row['cas'])} | {row['datum']} | {row['soutez']} | {row['liga']} | "
            f"{fmt_num(row['L'])} | {fmt_num(row['P'])} | {row['poradi'] or ''} |"
        )

    report_lines += [
        "",
        "## Top sestřiky",
        "",
        "| Hadice | Strana | Sestřik | Datum | Soutěž | Liga |",
        "|---|---|---:|---|---|---|",
    ]
    for row in top_target_rows[:20]:
        report_lines.append(
            f"| {row['hadice']} | {row['strana']} | {fmt_num(row['sestrik'])} | {row['datum']} | {row['soutez']} | "
            f"{row['liga']} |"
        )

    report_lines += [
        "",
        "## Ligy",
        "",
        "| Liga | Hadice | Počet | Nejlepší čas | Průměr výsledku | Medián výsledku | Umístění |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for row in [r for r in league_rows if r["hadice"] == "Celkem"]:
        report_lines.append(
            f"| {row['liga']} | {row['hadice']} | {row['pocet_zavodu']} | {fmt_num(row['nejlepsi_cas'])} | "
            f"{fmt_num(row['prumer_vysledku'])} | {fmt_num(row['median_vysledku'])} | {row['umisteni']} |"
        )

    (OUT_DIR / "statistiky_olesnice_x.md").write_text("\n".join(report_lines), encoding="utf-8")

    json_payload = {
            "source": str(source),
        "records": len(runs),
        "seasons": seasons,
        "overall": overall_rows,
    }
    (OUT_DIR / "statistiky_olesnice_x.json").write_text(json.dumps(json_payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps({"records": len(runs), "seasons": seasons, "output_dir": str(OUT_DIR)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
